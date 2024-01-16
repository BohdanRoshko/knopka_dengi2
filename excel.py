from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl import Workbook, load_workbook
from datetime import datetime
from time import sleep

from wallet import Wallet
from retry import logger


class Excel:
    def __init__(self, total_len: int):
        workbook = Workbook()
        sheet = workbook.active
        self.file_name = f'{total_len}accs_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'

        sheet['A1'] = 'access key'
        
        for cell in sheet._cells:
            sheet.cell(cell[0], cell[1]).font = Font(bold=True)

        sheet.column_dimensions['A'].width = 15
        

        for cell in ["A1", "B1", "C1"]:
            sheet[cell].alignment = Alignment(horizontal='center')

        workbook.save('results/'+self.file_name)


    def add_account(self, wallet: Wallet):
        while True:
            try:
                workbook = load_workbook('results/'+self.file_name)
                sheet = workbook.active

                max_row = sheet.max_row + 1

                # valid_info = [
                #     wallet.privatekey,
                #     wallet.address,
                #     str(wallet.status),
                # ]
                sheet.append([wallet])
                #if wallet.status != True: sheet.cell(max_row, sheet.max_column).fill = PatternFill(patternType='solid', fgColor=Color(rgb='ff0f0f'))

                workbook.save('results/'+self.file_name)
                return True
            except PermissionError:
                logger.warning(f'Excel | Cant save excel file, close it!')
                sleep(3)
            except Exception as err:
                logger.error(f'Excel | Cant save excel file: {err} | {wallet}')
                return False

